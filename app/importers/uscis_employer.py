from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import BinaryIO, Callable, TextIO

from sqlalchemy import select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.db import base as _models  # noqa: F401
from app.db.session import SessionLocal
from app.models.company import Company, CompanyAlias, SponsorshipTier
from app.models.uscis import UscisEmployerYearlyStat
from app.services.companies import slugify_company


def _key(value) -> str:
    return " ".join(str(value or "").replace("\ufeff", "").replace("\x00", "").split()).strip().lower()


HEADER_MAP = {
    key: _key(label)
    for key, label in {
        "fiscal_year": "Fiscal Year",
        "employer_name": "Employer (Petitioner) Name",
        "tax_id": "Tax ID",
        "naics": "Industry (NAICS) Code",
        "city": "Petitioner City",
        "state": "Petitioner State",
        "zip": "Petitioner Zip Code",
        "new_employment_approval": "New Employment Approval",
        "new_employment_denial": "New Employment Denial",
        "continuation_approval": "Continuation Approval",
        "continuation_denial": "Continuation Denial",
        "change_same_employer_approval": "Change with Same Employer Approval",
        "change_same_employer_denial": "Change with Same Employer Denial",
        "new_concurrent_approval": "New Concurrent Approval",
        "new_concurrent_denial": "New Concurrent Denial",
        "change_employer_approval": "Change of Employer Approval",
        "change_employer_denial": "Change of Employer Denial",
        "amended_approval": "Amended Approval",
        "amended_denial": "Amended Denial",
    }.items()
}


def import_uscis_employer_file(db: Session, path: str | Path, default_year: int | None = None, refresh: bool = True) -> dict[str, int]:
    file_path = Path(path)
    if file_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return import_uscis_employer_rows(db, _xlsx_rows(file_path), source_file=file_path.name, default_year=default_year, refresh=refresh)
    return import_uscis_employer_rows(db, _delimited_rows(file_path), source_file=file_path.name, default_year=default_year, refresh=refresh)


def bulk_import_uscis_employer_files(db: Session, paths: list[str | Path], default_year: int | None = None) -> dict[str, int]:
    company_names: dict[str, str] = {}
    aliases: dict[str, str] = {}
    stats: dict[tuple, dict] = {}
    skipped = 0
    source_rows = 0

    for path_value in paths:
        path = Path(path_value)
        rows = _xlsx_rows(path) if path.suffix.lower() in {".xlsx", ".xlsm"} else _delimited_rows(path)
        file_rows = 0
        file_skipped = 0
        for raw_row in rows:
            source_rows += 1
            file_rows += 1
            row = {_key(k): v for k, v in raw_row.items()}
            employer_name = str(row.get(HEADER_MAP["employer_name"]) or "").strip()
            if not employer_name:
                skipped += 1
                file_skipped += 1
                continue
            fiscal_year = _as_int(row.get(HEADER_MAP["fiscal_year"]) or default_year)
            if fiscal_year <= 0:
                skipped += 1
                file_skipped += 1
                continue

            normalized = normalize_uscis_employer_name(employer_name)
            company_names.setdefault(normalized, " ".join(employer_name.split()).title())
            aliases.setdefault(normalized, employer_name)
            naics_code, naics_label = _split_naics(row.get(HEADER_MAP["naics"]))
            identity = (
                fiscal_year,
                normalized,
                str(row.get(HEADER_MAP["tax_id"]) or "").strip(),
                naics_code,
                str(row.get(HEADER_MAP["city"]) or "").strip(),
                str(row.get(HEADER_MAP["state"]) or "").strip().upper(),
            )
            stat = stats.get(identity)
            if stat is None:
                stat = {
                    "fiscal_year": fiscal_year,
                    "normalized_employer_name": normalized,
                    "tax_id": identity[2],
                    "naics_code": naics_code,
                    "petitioner_city": identity[4],
                    "petitioner_state": identity[5],
                    "company_id": None,
                    "employer_name": employer_name,
                    "naics_label": naics_label,
                    "petitioner_zip_code": str(row.get(HEADER_MAP["zip"]) or "").strip(),
                    "new_employment_approval": 0,
                    "new_employment_denial": 0,
                    "continuation_approval": 0,
                    "continuation_denial": 0,
                    "change_same_employer_approval": 0,
                    "change_same_employer_denial": 0,
                    "new_concurrent_approval": 0,
                    "new_concurrent_denial": 0,
                    "change_employer_approval": 0,
                    "change_employer_denial": 0,
                    "amended_approval": 0,
                    "amended_denial": 0,
                    "total_approvals": 0,
                    "total_denials": 0,
                    "total_decisions": 0,
                    "source_file": path.name,
                }
                stats[identity] = stat
            stat["new_employment_approval"] += _as_int(row.get(HEADER_MAP["new_employment_approval"]))
            stat["new_employment_denial"] += _as_int(row.get(HEADER_MAP["new_employment_denial"]))
            stat["continuation_approval"] += _as_int(row.get(HEADER_MAP["continuation_approval"]))
            stat["continuation_denial"] += _as_int(row.get(HEADER_MAP["continuation_denial"]))
            stat["change_same_employer_approval"] += _as_int(row.get(HEADER_MAP["change_same_employer_approval"]))
            stat["change_same_employer_denial"] += _as_int(row.get(HEADER_MAP["change_same_employer_denial"]))
            stat["new_concurrent_approval"] += _as_int(row.get(HEADER_MAP["new_concurrent_approval"]))
            stat["new_concurrent_denial"] += _as_int(row.get(HEADER_MAP["new_concurrent_denial"]))
            stat["change_employer_approval"] += _as_int(row.get(HEADER_MAP["change_employer_approval"]))
            stat["change_employer_denial"] += _as_int(row.get(HEADER_MAP["change_employer_denial"]))
            stat["amended_approval"] += _as_int(row.get(HEADER_MAP["amended_approval"]))
            stat["amended_denial"] += _as_int(row.get(HEADER_MAP["amended_denial"]))
        print(f"Parsed {path.name}: rows={file_rows} skipped={file_skipped}", flush=True)

    for stat in stats.values():
        stat["total_approvals"] = (
            stat["new_employment_approval"]
            + stat["continuation_approval"]
            + stat["change_same_employer_approval"]
            + stat["new_concurrent_approval"]
            + stat["change_employer_approval"]
            + stat["amended_approval"]
        )
        stat["total_denials"] = (
            stat["new_employment_denial"]
            + stat["continuation_denial"]
            + stat["change_same_employer_denial"]
            + stat["new_concurrent_denial"]
            + stat["change_employer_denial"]
            + stat["amended_denial"]
        )
        stat["total_decisions"] = stat["total_approvals"] + stat["total_denials"]

    company_rows = []
    used_slugs: set[str] = set()
    existing_slugs = set(db.scalars(select(Company.slug)).all())
    existing_names = set(db.scalars(select(Company.name)).all())
    for normalized, name in company_names.items():
        if name in existing_names:
            continue
        base_slug = slugify_company(name)
        slug = base_slug
        suffix = 2
        while slug in existing_slugs or slug in used_slugs:
            slug = f"{base_slug}-{suffix}"[:280]
            suffix += 1
        used_slugs.add(slug)
        company_rows.append({"name": name, "slug": slug})

    if company_rows:
        _chunked_execute(db, pg_insert(Company).on_conflict_do_nothing(index_elements=["name"]), company_rows)
        db.commit()

    company_by_name = {company.name: company.id for company in db.scalars(select(Company)).all()}
    alias_rows = []
    existing_raw_aliases = set(db.scalars(select(CompanyAlias.raw_name)).all())
    for normalized, raw_name in aliases.items():
        if raw_name in existing_raw_aliases:
            continue
        company_id = company_by_name.get(company_names[normalized])
        if company_id:
            alias_rows.append({"company_id": company_id, "raw_name": raw_name, "normalized_name": normalized, "source": "uscis"})
    if alias_rows:
        _chunked_execute(db, pg_insert(CompanyAlias).on_conflict_do_nothing(index_elements=["raw_name"]), alias_rows)
        db.commit()

    company_by_normalized = {
        normalized: company_by_name[name]
        for normalized, name in company_names.items()
        if name in company_by_name
    }
    stat_rows = []
    for stat in stats.values():
        stat["company_id"] = company_by_normalized.get(stat["normalized_employer_name"])
        stat_rows.append(stat)

    upsert = pg_insert(UscisEmployerYearlyStat)
    update_fields = {
        column.name: getattr(upsert.excluded, column.name)
        for column in UscisEmployerYearlyStat.__table__.columns
        if column.name not in {"id", "created_at"}
    }
    statement = upsert.on_conflict_do_update(
        constraint="uq_uscis_employer_yearly_identity",
        set_=update_fields,
    )
    _chunked_execute(db, statement, stat_rows)
    db.commit()
    refresh_companies_from_uscis(db)
    return {"source_rows": source_rows, "imported_or_updated_stats": len(stat_rows), "companies_seen": len(company_names), "skipped": skipped}


def _chunked_execute(db: Session, statement, rows: list[dict], size: int = 5000) -> None:
    for index in range(0, len(rows), size):
        db.execute(statement, rows[index : index + size])
        print(f"Wrote {min(index + size, len(rows))}/{len(rows)}", flush=True)


def import_uscis_employer_rows(
    db: Session,
    rows_or_file,
    source_file: str,
    default_year: int | None = None,
    refresh: bool = True,
    progress_callback: Callable[[dict[str, int]], None] | None = None,
) -> dict[str, int]:
    rows = _delimited_file_rows(rows_or_file) if hasattr(rows_or_file, "read") else rows_or_file
    imported = 0
    updated = 0
    skipped = 0
    processed = 0
    touched_companies: set[int] = set()
    company_cache: dict[str, Company] = {}
    stat_cache: dict[tuple, UscisEmployerYearlyStat] = {}

    for raw_row in rows:
        processed += 1
        row = {_key(k): v for k, v in raw_row.items()}
        employer_name = str(row.get(HEADER_MAP["employer_name"]) or "").strip()
        if not employer_name:
            skipped += 1
            if progress_callback and processed % 2000 == 0:
                progress_callback({"processed_rows": processed, "imported": imported, "updated": updated, "skipped": skipped})
            continue

        fiscal_year = _as_int(row.get(HEADER_MAP["fiscal_year"]) or default_year)
        if fiscal_year <= 0:
            skipped += 1
            if progress_callback and processed % 2000 == 0:
                progress_callback({"processed_rows": processed, "imported": imported, "updated": updated, "skipped": skipped})
            continue

        normalized = normalize_uscis_employer_name(employer_name)
        company = _company_for_uscis_employer(db, employer_name, normalized, company_cache)
        naics_code, naics_label = _split_naics(row.get(HEADER_MAP["naics"]))
        identity = {
            "fiscal_year": fiscal_year,
            "normalized_employer_name": normalized,
            "tax_id": str(row.get(HEADER_MAP["tax_id"]) or "").strip(),
            "naics_code": naics_code,
            "petitioner_city": str(row.get(HEADER_MAP["city"]) or "").strip(),
            "petitioner_state": str(row.get(HEADER_MAP["state"]) or "").strip().upper(),
        }
        identity_key = tuple(identity.values())
        stat = stat_cache.get(identity_key)
        first_identity_seen = stat is None
        if stat is None:
            stat = db.scalar(
                select(UscisEmployerYearlyStat).where(
                    UscisEmployerYearlyStat.fiscal_year == identity["fiscal_year"],
                    UscisEmployerYearlyStat.normalized_employer_name == identity["normalized_employer_name"],
                    UscisEmployerYearlyStat.tax_id == identity["tax_id"],
                    UscisEmployerYearlyStat.naics_code == identity["naics_code"],
                    UscisEmployerYearlyStat.petitioner_city == identity["petitioner_city"],
                    UscisEmployerYearlyStat.petitioner_state == identity["petitioner_state"],
                )
            )
        if stat is None:
            stat = UscisEmployerYearlyStat(**identity)
            _reset_decision_counts(stat)
            imported += 1
        else:
            if first_identity_seen:
                updated += 1
                _reset_decision_counts(stat)
        stat_cache[identity_key] = stat

        stat.company_id = company.id
        stat.employer_name = employer_name
        stat.naics_label = naics_label
        stat.petitioner_zip_code = str(row.get(HEADER_MAP["zip"]) or "").strip()
        stat.new_employment_approval += _as_int(row.get(HEADER_MAP["new_employment_approval"]))
        stat.new_employment_denial += _as_int(row.get(HEADER_MAP["new_employment_denial"]))
        stat.continuation_approval += _as_int(row.get(HEADER_MAP["continuation_approval"]))
        stat.continuation_denial += _as_int(row.get(HEADER_MAP["continuation_denial"]))
        stat.change_same_employer_approval += _as_int(row.get(HEADER_MAP["change_same_employer_approval"]))
        stat.change_same_employer_denial += _as_int(row.get(HEADER_MAP["change_same_employer_denial"]))
        stat.new_concurrent_approval += _as_int(row.get(HEADER_MAP["new_concurrent_approval"]))
        stat.new_concurrent_denial += _as_int(row.get(HEADER_MAP["new_concurrent_denial"]))
        stat.change_employer_approval += _as_int(row.get(HEADER_MAP["change_employer_approval"]))
        stat.change_employer_denial += _as_int(row.get(HEADER_MAP["change_employer_denial"]))
        stat.amended_approval += _as_int(row.get(HEADER_MAP["amended_approval"]))
        stat.amended_denial += _as_int(row.get(HEADER_MAP["amended_denial"]))
        stat.total_approvals = (
            stat.new_employment_approval
            + stat.continuation_approval
            + stat.change_same_employer_approval
            + stat.new_concurrent_approval
            + stat.change_employer_approval
            + stat.amended_approval
        )
        stat.total_denials = (
            stat.new_employment_denial
            + stat.continuation_denial
            + stat.change_same_employer_denial
            + stat.new_concurrent_denial
            + stat.change_employer_denial
            + stat.amended_denial
        )
        stat.total_decisions = stat.total_approvals + stat.total_denials
        stat.source_file = source_file
        db.add(stat)
        touched_companies.add(company.id)

        if (imported + updated) % 2000 == 0:
            db.flush()
        if progress_callback and processed % 2000 == 0:
            progress_callback({"processed_rows": processed, "imported": imported, "updated": updated, "skipped": skipped})

    db.commit()
    if refresh:
        refresh_companies_from_uscis(db, touched_companies)
    if progress_callback:
        progress_callback({"processed_rows": processed, "imported": imported, "updated": updated, "skipped": skipped})
    return {"processed_rows": processed, "imported": imported, "updated": updated, "skipped": skipped}


def normalize_uscis_employer_name(name: str) -> str:
    value = (name or "").strip().lower()
    for marker in [" f/k/a ", " fka ", " formerly known as "]:
        if marker in value:
            value = value.split(marker, 1)[0]
    cleaned = [char if char.isalnum() else " " for char in value]
    suffixes = {"inc", "llc", "ltd", "corp", "corporation", "co", "company"}
    aliases = {"svcs": "services", "svc": "service"}
    tokens = [aliases.get(token, token) for token in "".join(cleaned).split()]
    while tokens and tokens[-1] in suffixes:
        tokens.pop()
    return " ".join(tokens) or value


def _reset_decision_counts(stat: UscisEmployerYearlyStat) -> None:
    stat.new_employment_approval = 0
    stat.new_employment_denial = 0
    stat.continuation_approval = 0
    stat.continuation_denial = 0
    stat.change_same_employer_approval = 0
    stat.change_same_employer_denial = 0
    stat.new_concurrent_approval = 0
    stat.new_concurrent_denial = 0
    stat.change_employer_approval = 0
    stat.change_employer_denial = 0
    stat.amended_approval = 0
    stat.amended_denial = 0
    stat.total_approvals = 0
    stat.total_denials = 0
    stat.total_decisions = 0


def _company_for_uscis_employer(db: Session, employer_name: str, normalized: str, cache: dict[str, Company]) -> Company:
    cached = cache.get(normalized)
    if cached:
        return cached

    alias = db.scalar(select(CompanyAlias).where(CompanyAlias.normalized_name == normalized).limit(1))
    if alias:
        cache[normalized] = alias.company
        return alias.company

    company_name = " ".join(employer_name.split()).title()
    existing_company = db.scalar(select(Company).where(Company.name == company_name).limit(1))
    if existing_company:
        alias = db.scalar(select(CompanyAlias).where(CompanyAlias.raw_name == employer_name).limit(1))
        if not alias:
            db.add(CompanyAlias(company_id=existing_company.id, raw_name=employer_name, normalized_name=normalized, source="uscis"))
        cache[normalized] = existing_company
        return existing_company

    base_slug = slugify_company(company_name)
    slug = base_slug
    suffix = 2
    while db.scalar(select(Company.id).where(Company.slug == slug)):
        slug = f"{base_slug}-{suffix}"[:280]
        suffix += 1
    company = Company(name=company_name, slug=slug)
    db.add(company)
    db.flush()
    db.add(CompanyAlias(company_id=company.id, raw_name=employer_name, normalized_name=normalized, source="uscis"))
    cache[normalized] = company
    return company


def refresh_companies_from_uscis(db: Session, company_ids: set[int] | None = None) -> None:
    company_filter = ""
    params = {}
    if company_ids:
        company_filter = "AND company_id = ANY(:company_ids)"
        params["company_ids"] = list(company_ids)
    db.execute(
        text(
            f"""
            WITH filtered AS (
                SELECT *
                FROM uscis_employer_yearly_stats
                WHERE company_id IS NOT NULL
                {company_filter}
            ),
            agg AS (
                SELECT
                    company_id,
                    SUM(total_approvals)::integer AS approvals,
                    SUM(total_denials)::integer AS denials,
                    SUM(total_decisions)::integer AS decisions,
                    SUM(new_employment_approval)::integer AS new_employment_approvals
                FROM filtered
                GROUP BY company_id
            ),
            latest_year AS (
                SELECT company_id, MAX(fiscal_year) AS fiscal_year
                FROM filtered
                GROUP BY company_id
            ),
            latest AS (
                SELECT
                    f.company_id,
                    SUM(f.total_decisions)::integer AS recent_filings,
                    SUM(f.new_employment_approval)::integer AS recent_new_employment_approvals
                FROM filtered f
                JOIN latest_year y ON y.company_id = f.company_id AND y.fiscal_year = f.fiscal_year
                GROUP BY f.company_id
            ),
            location_rank AS (
                SELECT
                    f.company_id,
                    f.petitioner_city,
                    f.petitioner_state,
                    ROW_NUMBER() OVER (
                        PARTITION BY f.company_id
                        ORDER BY SUM(f.total_approvals) DESC, SUM(f.total_decisions) DESC, f.petitioner_city, f.petitioner_state
                    ) AS rn
                FROM filtered f
                JOIN latest_year y ON y.company_id = f.company_id AND y.fiscal_year = f.fiscal_year
                WHERE COALESCE(NULLIF(f.petitioner_city, ''), '') <> ''
                   OR COALESCE(NULLIF(f.petitioner_state, ''), '') <> ''
                GROUP BY f.company_id, f.petitioner_city, f.petitioner_state
            ),
            industry_rank AS (
                SELECT
                    f.company_id,
                    f.naics_label,
                    ROW_NUMBER() OVER (
                        PARTITION BY f.company_id
                        ORDER BY SUM(f.total_approvals) DESC, SUM(f.total_decisions) DESC, f.naics_label
                    ) AS rn
                FROM filtered f
                JOIN latest_year y ON y.company_id = f.company_id AND y.fiscal_year = f.fiscal_year
                WHERE COALESCE(NULLIF(f.naics_label, ''), '') <> ''
                GROUP BY f.company_id, f.naics_label
            ),
            scored AS (
                SELECT
                    company_id,
                    approvals,
                    denials,
                    decisions,
                    new_employment_approvals,
                    CASE
                        WHEN decisions = 0 THEN 0
                        ELSE ROUND((approvals::numeric / decisions::numeric) * 100, 2)
                    END AS score
                FROM agg
            )
            UPDATE companies
            SET
                h1b_approval_count = scored.approvals,
                h1b_denial_count = scored.denials,
                h1b_filings_recent = COALESCE(latest.recent_filings, 0),
                opt_recent_hires = COALESCE(latest.recent_new_employment_approvals, 0),
                opt_last_verified = CURRENT_DATE,
                opt_friendliness_score = scored.score,
                opt_friendly = scored.approvals >= 3 AND scored.score >= 60,
                opt_status = CASE
                    WHEN companies.opt_status = 'unknown' AND COALESCE(latest.recent_new_employment_approvals, 0) > 0 THEN 'yes'
                    ELSE companies.opt_status
                END,
                sponsorship_status = CASE
                    WHEN companies.sponsorship_status = 'unknown' AND scored.approvals > 0 THEN 'yes'
                    ELSE companies.sponsorship_status
                END,
                opt_risk = CASE
                    WHEN companies.opt_risk IN ('', 'low') AND scored.approvals >= 5 AND scored.score >= 60 THEN 'low'
                    WHEN companies.opt_risk IN ('', 'low') AND scored.approvals > 0 THEN 'medium'
                    ELSE companies.opt_risk
                END,
                industry = COALESCE(industry_rank.naics_label, companies.industry),
                location = CASE
                    WHEN location_rank.rn = 1 THEN CONCAT_WS(', ', NULLIF(location_rank.petitioner_city, ''), NULLIF(location_rank.petitioner_state, ''), 'USA')
                    ELSE companies.location
                END,
                headquarters_city = COALESCE(location_rank.petitioner_city, companies.headquarters_city),
                headquarters_state = COALESCE(location_rank.petitioner_state, companies.headquarters_state),
                opt_notes = CASE
                    WHEN COALESCE(NULLIF(companies.opt_notes, ''), '') = '' THEN CONCAT(
                        'USCIS source summary: ',
                        scored.approvals, ' approvals, ',
                        scored.denials, ' denials, ',
                        scored.score, '% approval rate. Latest USCIS fiscal year: ',
                        latest_year.fiscal_year,
                        '; latest-year filings: ',
                        COALESCE(latest.recent_filings, 0),
                        '; latest-year new-employment approvals: ',
                        COALESCE(latest.recent_new_employment_approvals, 0),
                        CASE
                            WHEN location_rank.rn = 1 THEN CONCAT('; strongest filing location: ', CONCAT_WS(', ', NULLIF(location_rank.petitioner_city, ''), NULLIF(location_rank.petitioner_state, '')))
                            ELSE ''
                        END,
                        CASE
                            WHEN industry_rank.rn = 1 THEN CONCAT('; NAICS industry: ', industry_rank.naics_label)
                            ELSE ''
                        END,
                        '.'
                    )
                    ELSE companies.opt_notes
                END,
                sponsorship_tier = CASE
                    WHEN scored.approvals >= 25 AND scored.score >= 80 THEN 'high'
                    WHEN scored.approvals >= 5 AND scored.score >= 60 THEN 'medium'
                    WHEN scored.approvals > 0 THEN 'low'
                    ELSE 'unknown'
                END,
                updated_at = now()
            FROM scored
            LEFT JOIN latest_year ON latest_year.company_id = scored.company_id
            LEFT JOIN latest ON latest.company_id = scored.company_id
            LEFT JOIN location_rank ON location_rank.company_id = scored.company_id AND location_rank.rn = 1
            LEFT JOIN industry_rank ON industry_rank.company_id = scored.company_id AND industry_rank.rn = 1
            WHERE companies.id = scored.company_id
            """
        ),
        params,
    )
    db.commit()


def _delimited_rows(path: Path):
    encodings = ["utf-16", "utf-8-sig", "latin-1"]
    last_error = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                delimiter = "\t" if sample.count("\t") >= sample.count(",") else ","
                yield from csv.DictReader(handle, delimiter=delimiter)
                return
        except UnicodeError as exc:
            last_error = exc
            continue
    raise ValueError(f"Unable to decode {path}: {last_error}")


def _delimited_file_rows(file_obj: BinaryIO | TextIO):
    sample = file_obj.read(4096)
    if isinstance(sample, bytes):
        sample_text = sample.decode("utf-16", errors="ignore") if sample.startswith((b"\xff\xfe", b"\xfe\xff")) else sample.decode("utf-8-sig", errors="ignore")
    else:
        sample_text = sample
    file_obj.seek(0)
    delimiter = "\t" if sample_text.count("\t") >= sample_text.count(",") else ","
    return csv.DictReader(_text_lines(file_obj), delimiter=delimiter)


def _xlsx_rows(path: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return
    for values in iterator:
        yield {headers[index]: value for index, value in enumerate(values) if index < len(headers)}


def _text_lines(file_obj: BinaryIO | TextIO):
    for line in file_obj:
        if isinstance(line, bytes):
            if line.startswith((b"\xff\xfe", b"\xfe\xff")):
                yield line.decode("utf-16", errors="ignore")
            else:
                yield line.decode("utf-8-sig", errors="ignore")
        else:
            yield line


def _as_int(value) -> int:
    raw = str(value or "").strip().replace(",", "")
    if not raw:
        return 0
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return 0


def _split_naics(value) -> tuple[str, str]:
    raw = str(value or "").strip()
    if not raw:
        return "", ""
    if " - " in raw:
        code, label = raw.split(" - ", 1)
        return code.strip(), label.strip()
    return raw, ""


def main() -> None:
    parser = argparse.ArgumentParser(description="Import USCIS employer information files into Mintel.")
    parser.add_argument("paths", nargs="+")
    parser.add_argument("--default-year", type=int, default=None)
    args = parser.parse_args()
    totals = {"imported": 0, "updated": 0, "skipped": 0}
    with SessionLocal() as db:
        result = bulk_import_uscis_employer_files(db, args.paths, default_year=args.default_year)
    print(f"TOTAL: {result}")


if __name__ == "__main__":
    main()
